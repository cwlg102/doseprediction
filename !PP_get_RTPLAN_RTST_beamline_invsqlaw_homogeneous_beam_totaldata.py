import os 
import pydicom 
import numpy as np
import matplotlib.pyplot as plt 
import SimpleITK as sitk
import openpyxl 
from scipy.spatial.transform import Rotation as R
from shapely import Polygon
from multiprocessing import Pool
from skimage.draw import polygon, line_nd
from scipy.ndimage import rotate, binary_erosion, binary_dilation, binary_closing
from scipy.interpolate import interp1d
from PIL import Image 

def registration_for_mr(fixed_sitk, moving_sitk, min_val, param_save_path = None, ind = 0):
    """1) Set ElastixImageFilter"""
    elastixImageFilter = sitk.ElastixImageFilter()
    """2) Set Parameters"""
    elastixImageFilter.SetFixedImage(fixed_sitk)
    elastixImageFilter.SetMovingImage(moving_sitk)
    
    
    parameterMapVector = sitk.VectorOfParameterMap()
    translation_map = sitk.GetDefaultParameterMap('translation')
    translation_map['MaximumNumberOfIterations'] = ['500']
    translation_map['DefaultPixelValue'] = [str(min_val)]
    translation_map["Registration"] = ["MultiMetricMultiResolutionRegistration"]
    translation_map["FixedImagePyramid"] = ["FixedRecursiveImagePyramid"]
    translation_map["MovingImagePyramid"] = ["MovingRecursiveImagePyramid"]
    translation_map["ImagePyramidSchedule"] = ["8", "8" "4", "4", "2", "2", "1", "1"]
    translation_map["Interpolator"] = ["LinearInterpolator"]
    # translation_map["Metric"] = ["AdvancedNormalizedCorrelation"]
    translation_map["Optimizer"] = ["StandardGradientDescent"]
    translation_map["SP_a"] = ["24000"]
    translation_map['Metric1Weight'] = ["1e-4"]
    translation_map["Transform"] = ["TranslationTransform"]
    translation_map["AutomaticTransformInitialization"] = ["true"]
    translation_map["AutomaticScalesEstimation"] = ["true"]
    translation_map["FinalGridSpacingInPhysicalUnits"] = ["8.0", "4.0", "8.0"]
    translation_map["GridSpacingSchedule"] = ["8.0", "4.0", "2.5", "1.0"]
    translation_map["HowToCombineTransforms"] = ["Compose"]
    translation_map["AutomaticParameterEstimation"] = ["true"]
    translation_map["UseAdaptiveStepSizes"] = ["true"]
    translation_map["NumberOfHistogramBins"] = ["32"]
    translation_map["FixedKernelBSplineOrder"] = ["1"]
    translation_map["MovingKernelBSplineOrder"] = ["3"]
    translation_map["ImageSampler"] = ["RandomCoordinate"]
    translation_map["NumberOfSpatialSamples"] = ["8192"]
    translation_map["NewSamplesEveryIteration"] = ["true"]

    parameterMapVector.append(translation_map)
    # rigid_map = sitk.GetDefaultParameterMap("rigid")
    # # rigid_map["MaximumNumberOfSamplingAttempts"] = ["2"]
    # rigid_map['MaximumNumberOfIterations'] = ["3000"]
    # rigid_map['DefaultPixelValue'] = [str(min_val)]
    # parameterMapVector.append(rigid_map)
    bspline_map = sitk.GetDefaultParameterMap("bspline")
    bspline_map["UseDirectionCosines"] = ["true"]
    bspline_map["FixedImagePyramid"] = ["FixedRecursiveImagePyramid"]
    bspline_map["MovingImagePyramid"] = ["MovingRecursiveImagePyramid"]
    bspline_map["ImagePyramidSchedule"] = ["8", "8", "4", "4", "2", "2", "1", "1"]
    bspline_map['DefaultPixelValue'] = [str(min_val)]
    # bspline_map["ImagePyramidSchedule"] = ["8 8 4 4 2 2 1 1"]
    bspline_map["Interpolator"] = ["BSplineInterpolator"]
    bspline_map["Metric"] = ["AdvancedMattesMutualInformation", "TransformBendingEnergyPenalty"]
    bspline_map["Optimizer"] = ["AdaptiveStochasticGradientDescent"]
    bspline_map["FinalGridSpacingInPhysicalUnits"] = ["80.0",  "80.0", "40.0"]
    bspline_map['MaximumNumberOfIterations'] = ['120']
    bspline_map["AutomaticTransformInitialization"] = ["true"]
    bspline_map["GridSpacingSchedule"] = ["8.0", "4.0", "2.5", "1.0"]
    bspline_map["NumberOfSpatialSamples"] = ["2048"]
    bspline_map["NewSamplesEveryIteration"] = ["true"]
    bspline_map['DefaultPixelValue'] = [str(min_val)]
    # bspline_map["log_to_console"] = [""]
    # bspline_map["SP_a"] = ["50000"]
    bspline_map['Metric1Weight'] = ["5e-2"]
    parameterMapVector.append(bspline_map)
    
    # parametermap['MaximumNumberOfSamplingAttempts'] = ['16']
    elastixImageFilter.SetParameterMap(parameterMapVector) #method
    # elastixImageFilter.SetInitialTransform(initial_transform, inPlace=True)
    """3) Execute"""
    elastixImageFilter.Execute()
    resultimage = elastixImageFilter.GetResultImage()
    
    #save parameter map
    # temp_idx = 0
    # while True:
    #     try:
    #         sitk.WriteParameterFile(elastixImageFilter.GetTransformParameterMap()[temp_idx], os.path.join(param_save_path, "P%03d_%03d_param.txt" %(ind, temp_idx)))
    #     except:
    #         break
    #     temp_idx += 1

    return resultimage, elastixImageFilter.GetTransformParameterMap()
# from skimage.draw import polygon
def HistogramNorm(mr_volume):
    mr_one_dim = np.ravel(mr_volume)
    print(np.max(mr_volume), np.min(mr_volume))
    if np.min(mr_volume) < 0:
        mr_volume += np.min(mr_volume)
    
    #####normalization#####
    max_pixel_value = 4000
    mrnorm_hyperparam = 1000
    counts_list, bin_locations, patches = plt.hist(mr_one_dim, max_pixel_value, (0, max_pixel_value))
    plt.ylim((0, 1e5))
    plt.xlim((0, 2500))
    plt.xlabel("Pixel Value")
    plt.ylabel("Number of Pixels")
    # plt.show()
    
    for idx_val in range(max_pixel_value-1, -1, -1):
        if counts_list[idx_val] > mrnorm_hyperparam:
            val_norm = idx_val+1
            break
    plt.close()
    mr_volume = np.where(mr_volume > val_norm, val_norm, mr_volume)
    return mr_volume

def resample(sitk_volume, new_spacing, new_size, default_value=0, is_label=False):
    """1) Create resampler"""
    resample = sitk.ResampleImageFilter() 
    
    """2) Set parameters"""
    #set interpolation method, output direction, default pixel value
    resample.SetInterpolator(sitk.sitkLinear)
    resample.SetOutputDirection(sitk_volume.GetDirection())
    resample.SetDefaultPixelValue(default_value)
    
    #set output spacing
    new_spacing = np.array(new_spacing)
    resample.SetOutputSpacing(new_spacing)
    
    #set output size and origin
    old_size = np.array(sitk_volume.GetSize())
    old_spacing = np.array(sitk_volume.GetSpacing())
    new_size_no_shift = np.int16(np.round(old_size*old_spacing/new_spacing))
    old_origin = np.array(sitk_volume.GetOrigin())
    
    shift_amount = np.int16(np.round((new_size_no_shift - new_size)/2))*new_spacing
    new_origin = old_origin + shift_amount
    
    new_size = [int(s) for s in new_size]
    resample.SetSize(new_size)
    resample.SetOutputOrigin(new_origin)
    if is_label:
        resample.SetInterpolator(sitk.sitkNearestNeighbor)
    else:
        pass


    """3) execute"""
    new_volume = resample.Execute(sitk_volume)
    return new_volume
def beam_ray_tracing(dset):

    # 에너지와 감쇠 계수를 분리
    

    # 선형 보간 함수 생성
    
    line_draw_arr, area_co, beam_source_ras, ptv_arr, MV, spacing = dset
    
    
    for point in area_co:
        # ratio = 2
        # vector = point - beam_source_ras
        # extended_point = beam_source_ras + ratio * vector
        # extended_point = np.int32(np.round(extended_point))
        line_ras_co = np.squeeze(np.array(line_nd(beam_source_ras, point)))
        
        ras_distance = np.float16(np.abs(beam_source_ras - point))
        ras_distance *= spacing
        xy_diagonal_distance = np.sqrt(ras_distance[0] **2 + ras_distance[1] ** 2)
        xyz_diagonal_distance = np.sqrt(xy_diagonal_distance ** 2 + ras_distance[2] ** 2) # unit : mm
        xyz_diagonal_distance /= 10 # unit : cm
        
        # # x = np.linspace(0, xyz_diagonal_distance, num=len(line_ras_co[0]))
        # x *= -1
        # https://www.researchgate.net/figure/Reference-values-for-the-linear-attenuation-coefficient-LAC-at-511-keV_tbl1_328399993
        # # y = np.float32(MV * np.exp(-0.0001 * x))
        

        
        # y = np.array(list(y[::-1]))
        y_valid_idx = np.squeeze(np.where(line_ras_co[1] >= 0))
        line_ras_co = np.array([line_ras_co[0][y_valid_idx],line_ras_co[1][y_valid_idx],line_ras_co[2][y_valid_idx]])
        # # y = y[y_valid_idx]
        y_valid_idx = np.squeeze(np.where(line_ras_co[1] < ptv_arr.shape[1]))
        line_ras_co = np.array([line_ras_co[0][y_valid_idx],line_ras_co[1][y_valid_idx],line_ras_co[2][y_valid_idx]])
        # # y = y[y_valid_idx]
        line_ras_co = tuple(line_ras_co)
        line_draw_arr[line_ras_co[2], line_ras_co[1], line_ras_co[0]] = MV
        
    return line_draw_arr

if __name__ == "__main__":
    original_path = r"G:\gangnam_data"
    fol_dirs_list = os.listdir(original_path)
    xl_path = r"G:\! project\2024- UnityDosePrediction\data_acquisition_chart.xlsx"
    wb = openpyxl.load_workbook(xl_path)
    ws = wb["Sheet1"]
    
    plan_rs_basepath = r"G:\! project\2024- UnityDosePrediction\data\dose_generation"
    plan_rs_dirs_list = os.listdir(plan_rs_basepath)
    
    for idx in range(100):
        if idx < 52:
            continue
        
        what = str(ws.cell(idx+1, 2).value)
        if what is not None:
            pass 
        else:
            continue 

        MRN = str(ws.cell(idx+1, 1).value)
        
        
        
        dcm_path = os.path.join(original_path, MRN + "_all_fx", "1~MR1xT", "DCMData")
        dcm_dirs_list = os.listdir(dcm_path)
        mr_dcm_list = []
        for i in range(len(dcm_dirs_list)):
            ds = pydicom.dcmread(os.path.join(dcm_path, dcm_dirs_list[i]), force=True)
            ds.file_meta.TransferSyntaxUID = pydicom.uid.ImplicitVRLittleEndian
            IPPZ = float(ds.ImagePositionPatient[2])
            mr_dcm_list.append((ds, IPPZ))
        mr_dcm_list.sort(key = lambda x : x[1])

        mr_pix_list = []
        for i in range(len(mr_dcm_list)):
            ds = mr_dcm_list[i][0]
            mr_pix_list.append(ds.pixel_array * float(ds.RescaleSlope) + float(ds.RescaleIntercept))
        mr_arr = np.array(mr_pix_list, dtype=np.int32)
        mr_arr = HistogramNorm(mr_arr)
        
        ct_dcm_path = os.path.join(original_path, MRN + "_all_fx", "1~CT1", "DCMData")
        ct_dcm_dirs_list = os.listdir(ct_dcm_path)
        ct_dcm_list = []
        for i in range(len(ct_dcm_dirs_list)):
            ds = pydicom.dcmread(os.path.join(ct_dcm_path, ct_dcm_dirs_list[i]), force=True)
            ds.file_meta.TransferSyntaxUID = pydicom.uid.ImplicitVRLittleEndian
            IPPZ = float(ds.ImagePositionPatient[2])
            ct_dcm_list.append((ds, IPPZ))
        ct_dcm_list.sort(key = lambda x : x[1])

        ct_pix_list = []
        for i in range(len(ct_dcm_list)):
            ds = ct_dcm_list[i][0]
            ct_pix_list.append(ds.pixel_array * float(ds.RescaleSlope) + float(ds.RescaleIntercept))
        ct_arr = np.array(ct_pix_list, dtype=np.int32)
        
        ct_origin = np.array(ct_dcm_list[0][0].ImagePositionPatient)
        ct_spacing = np.array([ct_dcm_list[0][0].PixelSpacing[0], ct_dcm_list[0][0].PixelSpacing[1], float((ct_dcm_list[-1][1] - ct_dcm_list[0][1])/(len(ct_dcm_list)- 1))])
        
        
        beam_arr = np.zeros_like(mr_arr)
        
        origin = np.array(mr_dcm_list[0][0].ImagePositionPatient)
        spacing = np.array([mr_dcm_list[0][0].PixelSpacing[0], mr_dcm_list[0][0].PixelSpacing[1], float((mr_dcm_list[-1][1] - mr_dcm_list[0][1])/(len(mr_dcm_list)- 1))])

        for i in range(len(plan_rs_dirs_list)):
            if MRN in plan_rs_dirs_list[i]:
                if "Dose" not in plan_rs_dirs_list[i] and "StrctrSets" not in plan_rs_dirs_list[i]:
                    plan_path = os.path.join(plan_rs_basepath, plan_rs_dirs_list[i])
                elif "Dose" in plan_rs_dirs_list[i]:
                    dose_path = os.path.join(plan_rs_basepath, plan_rs_dirs_list[i])
                elif "StrctrSets" in plan_rs_dirs_list[i]:
                    rs_path = os.path.join(plan_rs_basepath, plan_rs_dirs_list[i])
                print(plan_rs_dirs_list[i])
        dose = pydicom.dcmread(dose_path, force=True)
        dose.file_meta.TransferSyntaxUID = pydicom.uid.ImplicitVRLittleEndian
        rtp = pydicom.dcmread(plan_path, force=True)
        rs = pydicom.dcmread(rs_path, force=True)

        ptv_arr = np.zeros_like(mr_arr)
        oar_dict = {}
        for rcs, ssr in zip(rs.ROIContourSequence, rs.StructureSetROISequence):
            if "ptv" in str(ssr.ROIName).lower() and "target" in str(ssr.ROIName).lower():
                print(ssr.ROIName)
                for ctr_data in rcs.ContourSequence:
                    co = np.array(ctr_data.ContourData)
                    co = np.reshape(co, (len(co)//3, 3))
                    co_raster = np.int32(np.round((co - origin)/(spacing)))
                    co_raster_x = co_raster[:, 0]
                    co_raster_y = co_raster[:, 1]
                    co_filled_y, co_filled_x = polygon(co_raster_y, co_raster_x)
                    ptv_arr[co_raster[0][2], co_filled_y, co_filled_x] = 1
                pass 
            else:
                if "Z" not in str(ssr.ROIName):
                    try:
                        ctr_arr = np.zeros_like(mr_arr)
                        for ctr_data in rcs.ContourSequence:
                            co = np.array(ctr_data.ContourData)
                            co = np.reshape(co, (len(co)//3, 3))
                            co_raster = np.int32(np.round((co - origin)/(spacing)))
                            co_raster_x = co_raster[:, 0]
                            co_raster_y = co_raster[:, 1]
                            co_filled_y, co_filled_x = polygon(co_raster_y, co_raster_x)
                            ctr_arr[co_raster[0][2], co_filled_y, co_filled_x] = 1
                        oar_dict[str(ssr.ROIName).lower()] = ctr_arr
                    except:
                        continue
                    
                else:
                    continue
                
        
            
            
            # co_list = []
            
            
                
                
                # co_list += list(co_filled)
            # co_raster_total = np.array(co_list)

        dose_arr = dose.pixel_array * float(dose.DoseGridScaling)
        dose_origin = np.array(dose.ImagePositionPatient)
        dose_spacing = np.array([3.0, 3.0, 3.0])
        dose_itk = sitk.GetImageFromArray(dose_arr)

        dose_itk.SetOrigin(dose_origin)
        dose_itk.SetSpacing(dose_spacing)

        ct_itk = sitk.GetImageFromArray(ct_arr)
        ct_itk.SetOrigin(ct_origin)
        ct_itk.SetSpacing(ct_spacing)
        
        mr_itk = sitk.GetImageFromArray(mr_arr)
        mr_itk.SetSpacing(spacing)
        mr_itk.SetOrigin(origin)
        ptv_itk = sitk.GetImageFromArray(ptv_arr)
        ptv_itk.SetSpacing(spacing)
        ptv_itk.SetOrigin(origin)
        

        dose_new_itk = sitk.Resample(dose_itk, mr_itk)

        # ct_itk = sitk.Resample(ct_itk, mr_itk)
        ct_itk, _ = registration_for_mr(mr_itk, ct_itk, -1024)
        
        sitk.WriteImage(ptv_itk, os.path.join(r"G:\! project\2024- UnityDosePrediction\data\processed_data\PTV", "P%03d_" %(idx) + MRN + "_ptv.nii.gz"))
        sitk.WriteImage(mr_itk, os.path.join(r"G:\! project\2024- UnityDosePrediction\data\processed_data\MR", "P%03d_" %(idx) + MRN + "_mr.nii.gz"))
        sitk.WriteImage(ct_itk, os.path.join(r"G:\! project\2024- UnityDosePrediction\data\processed_data\CT", "P%03d_" %(idx) + MRN + "_ct.nii.gz"))
        #######################################################################################################################################################
        sitk.WriteImage(dose_new_itk, os.path.join(r"G:\! project\2024- UnityDosePrediction\data\processed_data\DOSE", "P%03d_" %(idx) + MRN + "_dose.nii.gz"))
        ############################################################################################################################################################
        BEAM_SAVEPATH = r"G:\! project\2024- UnityDosePrediction\data\processed_data\BEAM"
        os.makedirs(os.path.join(BEAM_SAVEPATH, "P%03d_" %(idx) + MRN), exist_ok=True)
        OAR_SAVEPATH = r"G:\! project\2024- UnityDosePrediction\data\processed_data\OAR"
        os.makedirs(os.path.join(OAR_SAVEPATH, "P%03d_" %(idx) + MRN), exist_ok=True)
        for key,val in oar_dict.items():
            oar_itk = sitk.GetImageFromArray(val)
            oar_itk.SetOrigin(origin)
            oar_itk.SetSpacing(spacing)
            sitk.WriteImage(oar_itk, os.path.join(OAR_SAVEPATH, "P%03d_" %(idx) + MRN, "P%03d_" %(idx) + MRN + "_%s.nii.gz" %(key)))
        ptv_original_arr = np.copy(ptv_arr)
        
        import time

        

        for i in range(len(rtp.BeamSequence)):
            
            gantry_angle = rtp.BeamSequence[i].ControlPointSequence[0].GantryAngle
            ptv_arr = binary_dilation(ptv_original_arr, iterations=5)
            
            arr_y_size = ptv_arr.shape[1]
            arr_x_size = ptv_arr.shape[2]
            
            ptv_arr = np.pad(ptv_arr, ((0, 0), (180, 180), (180, 180)), mode= "constant", constant_values=0)
            start_time = time.time()
            
            ptv_arr = np.round(rotate(ptv_arr, gantry_angle, axes=(1, 2), reshape=False)).astype("uint8")
            
            ptv_line_arr = ptv_arr - binary_erosion(ptv_arr, iterations=1)*1
            # print(rtp.BeamSequence[i])
            # print(rtp.BeamSequence[i].ControlPointSequence[0])
            
            
            MV = float(rtp.BeamSequence[i].ControlPointSequence[0].NominalBeamEnergy)
            collimator_angle = rtp.BeamSequence[i].ControlPointSequence[0].BeamLimitingDeviceAngle
            patient_support_angle = rtp.BeamSequence[i].ControlPointSequence[0].PatientSupportAngle
            isocenter = np.array(rtp.BeamSequence[i].ControlPointSequence[0].IsocenterPosition)
            x_iso, y_iso, z_iso = isocenter
            SAD = rtp.BeamSequence[i].SourceAxisDistance
            SCD = rtp.BeamSequence[i].BeamLimitingDeviceSequence[1].SourceToBeamLimitingDeviceDistance

            pix_iso = np.int32((isocenter - origin)/spacing)
            
            

            x_beam = x_iso #+ SAD * np.sin(np.deg2rad(gantry_angle)) # beam 시작 x 좌표 
            y_beam = y_iso - SAD #* np.cos(np.deg2rad(gantry_angle)) # beam 시작 y 좌표 
            z_beam = z_iso
            
            beam_source_xyz = np.array([[x_beam, y_beam, z_beam]])
            beam_source_ras = np.int32(np.round((beam_source_xyz - origin)/ spacing))
            


            
            big_area_y = 0
            area_max = 0
            
            for y in range(ptv_arr.shape[1]):
                ptv_slice = ptv_arr[:, y, :]
                area = np.sum(ptv_slice)
                if area > area_max:
                    area_max = area
                    big_area_y = y 
            big_area_co = np.where(ptv_arr[:, big_area_y, :] == 1)
            big_area_z_co = big_area_co[0]
            big_area_x_co = big_area_co[1]
            big_area_y_co = np.zeros_like(big_area_co[0]) + big_area_y
            
            
            
            area_co = np.array([big_area_x_co, big_area_y_co, big_area_z_co])
            area_co = np.transpose(area_co)
            
            beam_source_ras = np.squeeze(beam_source_ras)
            # 배열 상에서 y 끝부분 
            ratio = (ptv_arr.shape[1]-1 - beam_source_ras[1]) / (area_co[0][1] - beam_source_ras[1])
            
            
            vector = area_co - beam_source_ras
            area_co = beam_source_ras + ratio * vector
            area_co = np.int32(np.round(area_co))
            area_co_x = area_co[:, 0]
            area_co_z = area_co[:, 2]
            area_co_x, area_co_z = polygon(area_co_x, area_co_z)
            area_co_y = np.zeros_like(area_co_x) + area_co[0][1]
            
            temp_arr = np.zeros_like(ptv_arr).astype("uint8")
            temp_arr[area_co_z, area_co_y, area_co_x] = 1
            
            
            temp_arr[:, temp_arr.shape[1]-1, :] = np.uint8(binary_closing(temp_arr[:, temp_arr.shape[1]-1, :], iterations=1))        
            
            area_co = np.transpose(np.array(np.where(temp_arr == 1)))
            area_co = np.flip(area_co, axis = 1)
            line_draw_arr = np.zeros_like(ptv_arr).astype("float32")
            
            print("line drawing...")

            data_split = []
            for split_idx in range(4):
                data_split.append(
                    [
                    line_draw_arr,
                    area_co[split_idx * len(area_co)//4: (split_idx + 1) * len(area_co)//4],
                    beam_source_ras,
                    ptv_arr,
                    MV,
                    spacing
                    ])
                
            # area_co_1 = area_co[:len(area_co)//4]
            # area_co_2 = area_co[len(area_co)//4: 2 * len(area_co)//4]
            # area_co_3 = area_co[2 * len(area_co)//4: 3 * len(area_co)//4]
            # area_co_4 = area_co[3 * len(area_co)//4: 4 * len(area_co)//4]
            
            # data_split = [[line_draw_arr, area_co_1, beam_source_ras, ptv_arr],
            #               [line_draw_arr, area_co_2, beam_source_ras, ptv_arr],
            #               [line_draw_arr, area_co_3, beam_source_ras, ptv_arr],
            #               [line_draw_arr, area_co_4, beam_source_ras, ptv_arr]]
            with Pool(processes=4) as pool:
                results = pool.map(beam_ray_tracing, data_split)
            line_draw_arr = np.squeeze(np.sum(np.array(results), axis = 0)).astype(np.float32)
            
            # weight_arr = np.zeros_like(ptv_arr).astype("uint8")
            # for temp in results:
            #     weight_arr += np.where(temp >= 0.01, 1, 0).astype("uint8")
            # weight_co = np.where(weight_arr == 2)
            # line_draw_arr[weight_co[0], weight_co[1], weight_co[2]] /= 2
            
            line_draw_arr[line_draw_arr > MV] = MV

            line_draw_arr = np.round(rotate(line_draw_arr, -1 * gantry_angle, axes=(1, 2), reshape=False)).astype(np.uint8)
            line_draw_arr = line_draw_arr[:, 180:arr_y_size + 180, 180:arr_x_size+180]
            # line_draw_arr[:, 300:ptv_line_arr.shape[1]-300, 300:ptv_line_arr.shape[2]-300]
            # plt.imshow(line_draw_arr[125])
            
            
            
            line_itk = sitk.GetImageFromArray(line_draw_arr)
            line_itk.SetSpacing(mr_itk.GetSpacing())
            line_itk.SetOrigin(mr_itk.GetOrigin())
            sitk.WriteImage(line_itk, os.path.join(BEAM_SAVEPATH, "P%03d_" %(idx)+MRN, "P%03d_"%(idx) + MRN +"_beamline_%03d.nii.gz" %(int(gantry_angle))))
        
        # break
    

    
        
    
    

    
    
# ct_img_arr = np.transpose(ct_img_arr, (1, 0, 2, 3)).astype(np.uint8)
# from PIL import Image 
# for i in range(len(ct_img_arr)):
#     # if i == 399:
#     #     pass 
#     # else:
#     #     continue
#     img = Image.fromarray(ct_img_arr[i])
#     img.save(os.path.join(r"G:\! project\2024- UnityDosePrediction\img", "s%03d.png" %(i)))